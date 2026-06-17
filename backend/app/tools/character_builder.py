from __future__ import annotations

import copy
from pathlib import Path

from openpyxl import load_workbook

from app.config import settings
from app.tools.character_rules import (
    ABILITY_KEYS,
    derive_character_rules,
    point_buy_cost,
    validate_character_rules,
)
from app.tools.spell_catalog import enrich_character_spells
from app.tools.item_schema import normalize_inventory, CurrencyWallet
from app.tools.effect_engine import normalize_effects

ABILITY_CELLS = {"str": "E8", "dex": "E10", "con": "E12", "int": "E14", "wis": "E16", "cha": "E18"}

# 技能名 -> (熟练标记列, 总值列) in 主要情况 sheet
SKILL_CELL_MAP: dict[str, tuple[str, str]] = {
    "athletics":        ("B23", "I23"),
    "acrobatics":       ("B25", "I25"),
    "sleight_of_hand":  ("B26", "I26"),
    "stealth":          ("B27", "I27"),
    "investigation":    ("B29", "I29"),
    "arcana":           ("B30", "I30"),
    "history":          ("B31", "I31"),
    "nature":           ("B32", "I32"),
    "religion":         ("B33", "I33"),
    "perception":       ("B35", "I35"),
    "insight":          ("B36", "I36"),
    "animal_handling":  ("B37", "I37"),
    "medicine":         ("B38", "I38"),
    "survival":         ("B39", "I39"),
    "persuasion":       ("B41", "I41"),
    "deception":        ("B42", "I42"),
    "intimidation":     ("B43", "I43"),
    "performance":      ("B44", "I44"),
}

# 武器列起始行（主要情况 sheet 24-30 行）
WEAPON_START_ROW = 24
MAX_WEAPON_ROWS = 7

# 特性列起始行
FEATURE_CLASS_START_ROW = 1  # AV1 header
FEATURE_RACE_START_ROW = 1   # BM1 header


def build_character_data(raw: dict) -> dict:
    abilities = {key: int(raw.get("abilities", {}).get(key, 10)) for key in ABILITY_KEYS}
    level = int(raw.get("level", 1))
    derived = derive_character_rules(raw)
    hit_die = derived["hit_die"]
    max_hp = int(raw.get("max_hp") or hit_die + derived["ability_modifiers"]["con"])
    armor_class = int(raw.get("armor_class") or 10 + derived["ability_modifiers"]["dex"])
    return {
        "basic": {
            "name": raw["character_name"],
            "actor_type": raw.get("actor_type") if raw.get("actor_type") in {"npc", "monster"} else "player",
            "ancestry": raw.get("ancestry", ""),
            "subrace": raw.get("subrace", ""),
            "background": raw.get("background", ""),
            "alignment": raw.get("alignment", ""),
            "classes": [{"name": raw.get("class_name", ""), "level": level, "hit_die": hit_die}],
            "gender": raw.get("gender", ""),
            "age": raw.get("age", ""),
            "faith": raw.get("faith", ""),
            "appearance": raw.get("appearance", ""),
            "hair": raw.get("hair", ""),
            "height": raw.get("height", ""),
            "skin": raw.get("skin", ""),
            "weight": raw.get("weight", ""),
            "eyes": raw.get("eyes", ""),
        },
        "abilities": abilities,
        "combat": {
            "armor_class": armor_class,
            "max_hp": max_hp,
            "current_hp": max_hp,
            "temp_hp": 0,
            "proficiency_bonus": derived["proficiency_bonus"],
            "speed": int(raw.get("speed", 30)),
            "hit_dice": {"die": f"d{hit_die}", "maximum": level, "current": level},
            "initiative": derived["initiative"],
            "passive_perception": derived["passive_perception"],
            "carrying_capacity": derived["carrying_capacity"],
        },
        "saving_throw_proficiencies": derived["saving_throw_proficiencies"],
        "saving_throws": derived["saving_throws"],
        "skills": derived["skills"],
        "proficiencies": {
            "languages": list(raw.get("languages", [])),
            "tools": list(raw.get("tool_proficiencies", [])),
            "weapons": list(raw.get("weapon_proficiencies", [])),
            "armor": list(raw.get("armor_proficiencies", [])),
        },
        "inventory": normalize_inventory(raw.get("inventory", [])),
        "currency": CurrencyWallet.model_validate(raw.get("currency") or {}).model_dump(mode="json"),
        "features": copy.deepcopy(raw.get("features", [])),
        "spells": enrich_character_spells(copy.deepcopy(raw.get("spells", [])), settings.data_dir),
        "spellcasting": {
            "ability": raw.get("spellcasting_ability", ""),
            "save_dc": derived["spell_save_dc"],
            "attack_bonus": derived["spell_attack_bonus"],
        },
        "derived": derived,
        "validation_errors": validate_character_rules(raw, derived),
        "personality": {
            "traits": raw.get("traits", ""),
            "ideals": raw.get("ideals", ""),
            "bonds": raw.get("bonds", ""),
            "flaws": raw.get("flaws", ""),
            "backstory": raw.get("backstory", ""),
        },
        "roleplay": copy.deepcopy(raw.get("roleplay", {})),
        "story_role": copy.deepcopy(raw.get("story_role", {})),
        "encounter": {"present": True, **copy.deepcopy(raw.get("encounter", {}))},
        "conditions": [],
        "active_effects": normalize_effects(raw.get("active_effects")),
        "notes": copy.deepcopy(raw.get("notes", {})),
        "integrations": {"qq_user_ids": []},
    }


def export_character_sheet(data: dict, player_name: str, template: Path, target: Path) -> Path:
    """将角色数据按 D&D 5E 人物卡模板导出为 XLSX。"""
    from app.tools.character_rules import ability_modifier

    workbook = load_workbook(template)
    identity = workbook.worksheets[0]       # 角色
    main_sheet = workbook.worksheets[1]     # 主要情况
    backpack = workbook.worksheets[2]        # 背包
    spellcasting = workbook.worksheets[3]   # 施法

    basic = data.get("basic", {})
    classes = basic.get("classes") or [{}]
    primary_class = classes[0]
    abilities = data.get("abilities", {})
    combat = data.get("combat", {})
    derived = data.get("derived", {})
    personality = data.get("personality", {})
    saving_throws = data.get("saving_throws", {})
    saving_profs = data.get("saving_throw_proficiencies", [])
    skills = data.get("skills", {})
    proficiencies = data.get("proficiencies", {})
    inventory_items = data.get("inventory", [])
    currency = data.get("currency", {})
    features = data.get("features", [])
    spells = data.get("spells", [])
    spellcasting_data = data.get("spellcasting", {})

    # ═══════════════════════════════════════
    # Sheet 0: 角色 (Identity)
    # ═══════════════════════════════════════
    identity_values = {
        "E3": basic.get("name", ""),
        "E4": player_name,
        "E6": basic.get("ancestry", ""),
        "M6": basic.get("alignment", ""),
        "E7": basic.get("gender", ""),
        "M7": basic.get("age", ""),
        "E8": basic.get("subrace", ""),
        "M8": basic.get("faith", ""),
        "E10": basic.get("background", ""),
        "J10": basic.get("hair", ""),
        "B11": basic.get("height", ""),
        "J11": basic.get("skin", ""),
        "B12": basic.get("weight", ""),
        "J12": basic.get("eyes", ""),
        "B15": basic.get("appearance", ""),
        "R15": personality.get("traits", ""),
        "R16": personality.get("ideals", ""),
        "R17": personality.get("bonds", ""),
        "R18": personality.get("flaws", ""),
        "R19": personality.get("backstory", ""),
    }
    for cell, value in identity_values.items():
        set_sheet_value(identity, cell, value)

    # ═══════════════════════════════════════
    # Sheet 1: 主要情况 (Main Stats)
    # ═══════════════════════════════════════
    set_sheet_value(main_sheet, "S3", primary_class.get("name", ""))
    set_sheet_value(main_sheet, "W3", int(primary_class.get("level", 1)))

    # 熟练加值
    pb = int(derived.get("proficiency_bonus", 2))
    set_sheet_value(main_sheet, "AC3", pb)

    ability_mods = derived.get("ability_modifiers", {})
    for key, cell in ABILITY_CELLS.items():
        score = int(abilities.get(key, 10))
        mod = int(ability_mods.get(key, ability_modifier(score)))
        set_sheet_value(main_sheet, cell, score)
        # G 列 = 调整值
        g_col = "G" + cell[1:]
        set_sheet_value(main_sheet, g_col, mod)

    # 豁免熟练标记 + 豁免值
    save_mark_cells = {"str": "B8", "dex": "B10", "con": "B12", "int": "B14", "wis": "B16", "cha": "B18"}
    save_total_cells = {"str": "K8", "dex": "K10", "con": "K12", "int": "K14", "wis": "K16", "cha": "K18"}
    for key in ABILITY_CELLS:
        prof_mark = "O" if key in saving_profs else "X"
        set_sheet_value(main_sheet, save_mark_cells[key], prof_mark)
        set_sheet_value(main_sheet, save_total_cells[key], int(saving_throws.get(key, 0)))

    # 战斗属性
    set_sheet_value(main_sheet, "P8", int(combat.get("initiative", 0)))          # 先攻
    set_sheet_value(main_sheet, "P10", int(combat.get("armor_class", 10)))       # AC
    set_sheet_value(main_sheet, "Y8", int(combat.get("current_hp", 0)))          # 当前HP
    set_sheet_value(main_sheet, "AC8", int(combat.get("max_hp", 1)))             # 最大HP
    set_sheet_value(main_sheet, "Y14", int(combat.get("temp_hp", 0)))            # 临时HP
    set_sheet_value(main_sheet, "AD17", int(combat.get("speed", 30)))            # 速度
    set_sheet_value(main_sheet, "P14", int(spellcasting_data.get("save_dc", 0) or 0))  # 法术DC
    set_sheet_value(main_sheet, "R19", int(combat.get("passive_perception", 10))) # 被动察觉

    # 生命骰
    hit_dice = combat.get("hit_dice", {})
    set_sheet_value(main_sheet, "AC10", int(hit_dice.get("maximum", 1)))
    set_sheet_value(main_sheet, "AB11", int(hit_dice.get("current", hit_dice.get("maximum", 1))))

    # 施法关键属性
    spell_ability = spellcasting_data.get("ability", "")
    set_sheet_value(main_sheet, "R17", spell_ability)

    # 技能: 熟练标记 + 总值
    for skill_key, (mark_cell, total_cell) in SKILL_CELL_MAP.items():
        skill_info = skills.get(skill_key, {})
        if not isinstance(skill_info, dict):
            skill_info = {}
        if skill_info.get("expertise"):
            mark = "❂"
        elif skill_info.get("proficient"):
            mark = "O"
        else:
            mark = "X"
        set_sheet_value(main_sheet, mark_cell, mark)
        set_sheet_value(main_sheet, total_cell, int(skill_info.get("bonus", 0)))

    # 熟练项概述
    prof_lines: list[str] = []
    for cat_key, cat_label in [("languages", "语言"), ("tools", "工具"), ("weapons", "武器"), ("armor", "护甲")]:
        items = proficiencies.get(cat_key, [])
        if items:
            prof_lines.append(f"{cat_label}: {', '.join(str(i) for i in items)}")
    if prof_lines:
        set_sheet_value(identity, "R3", "\n".join(prof_lines[:4]))

    # 武器（L24-AQ30 区域）
    equipped_weapons = [
        item for item in inventory_items
        if isinstance(item, dict) and item.get("item_type") == "weapon" and item.get("equipped")
    ]
    for idx, weapon in enumerate(equipped_weapons[:MAX_WEAPON_ROWS]):
        row = WEAPON_START_ROW + idx
        set_sheet_value(main_sheet, f"O{row}", weapon.get("name", ""))
        set_sheet_value(main_sheet, f"V{row}", ", ".join(weapon.get("properties", [])))
        set_sheet_value(main_sheet, f"AA{row}", weapon.get("weight", ""))
        set_sheet_value(main_sheet, f"AE{row}", f"d20+{weapon.get('attack_bonus', 0)}")
        set_sheet_value(main_sheet, f"AI{row}", weapon.get("damage", ""))
        set_sheet_value(main_sheet, f"AM{row}", weapon.get("damage_type", ""))

    # 护甲
    equipped_armor = next(
        (item for item in inventory_items
         if isinstance(item, dict) and item.get("item_type") == "armor" and item.get("equipped")),
        None,
    )
    if equipped_armor:
        set_sheet_value(main_sheet, "O22", equipped_armor.get("name", ""))
        set_sheet_value(main_sheet, "L23", equipped_armor.get("armor_type", ""))
        set_sheet_value(main_sheet, "AD23", int(equipped_armor.get("base_ac", 10)))

    # 特性（AV 列 = 职业特性, BM 列 = 种族特性）
    class_features = [f for f in features if isinstance(f, dict) and f.get("source") != "race"]
    race_features = [f for f in features if isinstance(f, dict) and f.get("source") == "race"]
    for idx, feat in enumerate(class_features[:10]):
        row = 15 + idx
        set_sheet_value(main_sheet, f"AV{row}", feat.get("name", ""))
        set_sheet_value(main_sheet, f"BB{row}", feat.get("description", ""))
    for idx, feat in enumerate(race_features[:10]):
        row = 15 + idx
        set_sheet_value(main_sheet, f"BM{row}", feat.get("name", ""))
        set_sheet_value(main_sheet, f"BS{row}", feat.get("description", ""))

    # ═══════════════════════════════════════
    # Sheet 2: 背包 (Backpack)
    # ═══════════════════════════════════════
    carried_items = [
        item for item in inventory_items
        if isinstance(item, dict) and not item.get("equipped") and item.get("item_type") != "weapon"
    ]
    carried_items += [
        item for item in inventory_items
        if isinstance(item, dict) and not item.get("equipped") and item.get("item_type") == "weapon"
    ]
    bp_start = 5  # 背包物品起始行（左列 5-13, 右列 5-13）
    for idx, item in enumerate(carried_items[:18]):
        if idx < 9:
            row = bp_start + idx
            name_col = "B"; desc_col = "I"; weight_col = "S"; qty_col = "V"
        else:
            row = bp_start + idx - 9  # 右列同样从 row 5 开始
            name_col = "Y"; desc_col = "AF"; weight_col = "AP"; qty_col = "AS"
        set_sheet_value(backpack, f"{name_col}{row}", item.get("name", ""))
        set_sheet_value(backpack, f"{desc_col}{row}", item.get("description", ""))
        set_sheet_value(backpack, f"{weight_col}{row}", item.get("weight", ""))
        set_sheet_value(backpack, f"{qty_col}{row}", item.get("quantity", 1))

    # 货币
    set_sheet_value(backpack, "J51", currency.get("cp", 0))
    set_sheet_value(backpack, "J52", currency.get("sp", 0))
    set_sheet_value(backpack, "J53", currency.get("ep", 0))
    set_sheet_value(backpack, "J54", currency.get("gp", 0))
    set_sheet_value(backpack, "J55", currency.get("pp", 0))

    # ═══════════════════════════════════════
    # Sheet 3: 施法 (Spellcasting)
    # ═══════════════════════════════════════
    set_sheet_value(spellcasting, "H2", spell_ability)                   # 施法关键属性
    set_sheet_value(spellcasting, "O2", spellcasting_data.get("save_dc", 0) or 0)   # 法术DC
    set_sheet_value(spellcasting, "V2", spellcasting_data.get("attack_bonus", 0) or 0)  # 法术命中

    # 法术位（AX 列 = 当前, BA 列 = 最大）
    spell_slots = data.get("spell_slots", {})
    for level in range(0, 9):
        row = 5 + level
        current = spell_slots.get(str(level), 0)
        maximum = spell_slots.get(f"{level}_max", current)
        set_sheet_value(spellcasting, f"AX{row}", int(current))
        set_sheet_value(spellcasting, f"BA{row}", int(maximum))

    # 已知法术
    spells_by_level: dict[int, list[dict]] = {}
    for sp in spells:
        if isinstance(sp, dict):
            lv = int(sp.get("level", 0))
            spells_by_level.setdefault(lv, []).append(sp)
    spell_row = 4
    for level in sorted(spells_by_level):
        for sp in spells_by_level[level]:
            set_sheet_value(spellcasting, f"D{spell_row}", sp.get("name", ""))
            set_sheet_value(spellcasting, f"E{spell_row}", sp.get("school", ""))
            set_sheet_value(spellcasting, f"H{spell_row}", "R" if sp.get("ritual") else "")
            set_sheet_value(spellcasting, f"P{spell_row}", sp.get("casting_time", ""))
            set_sheet_value(spellcasting, f"R{spell_row}", sp.get("range", ""))
            set_sheet_value(spellcasting, f"U{spell_row}", sp.get("duration", ""))
            set_sheet_value(spellcasting, f"AA{spell_row}", "Y" if sp.get("verbal") else "")
            set_sheet_value(spellcasting, f"AB{spell_row}", "Y" if sp.get("somatic") else "")
            set_sheet_value(spellcasting, f"AC{spell_row}", "Y" if sp.get("material") else "")
            set_sheet_value(spellcasting, f"AD{spell_row}", sp.get("components", ""))
            set_sheet_value(spellcasting, f"AL{spell_row}", sp.get("description", ""))
            set_sheet_value(spellcasting, f"AV{spell_row}", level)
            spell_row += 1

    workbook.save(target)
    return target


def set_sheet_value(worksheet, coordinate: str, value) -> None:
    cell = worksheet[coordinate]
    if cell.__class__.__name__ != "MergedCell":
        cell.value = value
        return
    for merged in worksheet.merged_cells.ranges:
        if coordinate in merged:
            worksheet.cell(merged.min_row, merged.min_col).value = value
            return
    raise ValueError(f"Unable to resolve merged cell {worksheet.title}!{coordinate}")
