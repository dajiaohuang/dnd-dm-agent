from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from zipfile import ZipFile

from lxml import etree
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator


def local(node) -> str:
    return etree.QName(node).localname


def direct(node, name: str):
    return next((child for child in node if local(child) == name), None)


def text_of(node) -> str:
    return "".join(node.itertext()) if node is not None else ""


def esc(value) -> str:
    text = "" if value is None else str(value)
    return text.replace("\\", "\\\\").replace("|", "\\|").replace("\r", "").replace("\n", "<br>")


def code(value) -> str:
    text = "" if value is None else str(value)
    fence = "```" if "```" not in text else "````"
    return f"{fence}\n{text}\n{fence}"


def shared_strings(zf: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = etree.fromstring(zf.read("xl/sharedStrings.xml"))
    return ["".join(item.itertext()) for item in root if local(item) == "si"]


def decode_value(cell, strings: list[str]) -> str:
    kind = cell.get("t", "n")
    if kind == "inlineStr":
        return text_of(direct(cell, "is"))
    raw = text_of(direct(cell, "v"))
    if kind == "s" and raw:
        try:
            return strings[int(raw)]
        except (ValueError, IndexError):
            return raw
    if kind == "b":
        return "TRUE" if raw == "1" else "FALSE"
    return raw


def cell_records(root, strings: list[str]) -> list[dict]:
    records = []
    masters: dict[str, tuple[str, str]] = {}
    for cell in root.xpath('//*[local-name()="c"]'):
        coord = cell.get("r", "?")
        formula_node = direct(cell, "f")
        formula = text_of(formula_node)
        if formula_node is not None and formula_node.get("t") == "shared" and formula:
            masters[formula_node.get("si", "")] = (coord, formula)
        records.append({
            "coord": coord,
            "type": cell.get("t", "n"),
            "style": cell.get("s", "0"),
            "formula_node": formula_node,
            "formula": formula,
            "cached": decode_value(cell, strings),
        })
    for item in records:
        node = item.pop("formula_node")
        if node is None:
            item["value"] = item.pop("cached")
            continue
        if node.get("t") == "shared" and not item["formula"]:
            master = masters.get(node.get("si", ""))
            if master:
                try:
                    item["formula"] = Translator("=" + master[1], origin=master[0]).translate_formula(item["coord"])[1:]
                except Exception:
                    item["formula"] = f"[shared si={node.get('si')}; master={master[0]}]"
        item["formula_type"] = node.get("t", "normal")
        item["formula_ref"] = node.get("ref", "")
    return records


def validation_records(root) -> list[dict]:
    result = []
    for node in root.xpath('//*[local-name()="dataValidation"]'):
        refs = node.get("sqref", "")
        if not refs:
            sqref = direct(node, "sqref")
            refs = text_of(sqref)
        formulas = []
        for child in node.iterdescendants():
            if local(child) in {"formula1", "formula2", "f"} and text_of(child):
                formulas.append(f"{local(child)}={text_of(child)}")
        result.append({
            "sqref": refs,
            "type": node.get("type", ""),
            "operator": node.get("operator", ""),
            "allowBlank": node.get("allowBlank", ""),
            "showErrorMessage": node.get("showErrorMessage", ""),
            "promptTitle": node.get("promptTitle", ""),
            "prompt": node.get("prompt", ""),
            "errorTitle": node.get("errorTitle", ""),
            "error": node.get("error", ""),
            "formulas": "; ".join(dict.fromkeys(formulas)),
        })
    return result


def conditional_records(root) -> list[dict]:
    result = []
    for block in root.xpath('//*[local-name()="conditionalFormatting"]'):
        sqref = block.get("sqref", "") or text_of(direct(block, "sqref"))
        for rule in block.xpath('.//*[local-name()="cfRule"]'):
            formulas = [text_of(x) for x in rule.xpath('.//*[local-name()="formula" or local-name()="f"]') if text_of(x)]
            result.append({
                "sqref": sqref,
                "type": rule.get("type", ""),
                "priority": rule.get("priority", ""),
                "operator": rule.get("operator", ""),
                "dxfId": rule.get("dxfId", ""),
                "formulas": "; ".join(formulas),
            })
    return result


def sheet_formula_stats(records: list[dict]) -> tuple[Counter, Counter]:
    functions = Counter()
    dependencies = Counter()
    for item in records:
        formula = item.get("formula", "")
        for name in re.findall(r"(?<![A-Z0-9_.])([A-Z][A-Z0-9_.]*)\s*\(", formula, flags=re.I):
            functions[name.upper()] += 1
        for quoted, plain in re.findall(r"(?:'([^']+)'|([\w\u4e00-\u9fff ]+))!", formula):
            dependencies[(quoted or plain).strip()] += 1
    return functions, dependencies


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    source = args.xlsx.resolve()
    output = args.output.resolve()
    wb = load_workbook(source, data_only=False)
    cached_wb = load_workbook(source, data_only=True)

    with ZipFile(source) as zf:
        strings = shared_strings(zf)
        workbook_root = etree.fromstring(zf.read("xl/workbook.xml"))
        rels_root = etree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {node.get("Id"): node.get("Target", "") for node in rels_root}
        raw_sheet_parts = []
        for node in workbook_root.xpath('//*[local-name()="sheets"]/*[local-name()="sheet"]'):
            rel_id = next((value for key, value in node.attrib.items() if key.endswith("}id") or key == "r:id"), "")
            target = rel_targets.get(rel_id, "")
            raw_sheet_parts.append(target.lstrip("/") if target.startswith("/") else "xl/" + target.lstrip("/"))
        sheet_data = []
        all_functions = Counter()
        all_dependencies: dict[str, Counter] = {}
        for index, ws in enumerate(wb.worksheets):
            part = raw_sheet_parts[index]
            root = etree.fromstring(zf.read(part))
            records = cell_records(root, strings)
            validations = validation_records(root)
            conditionals = conditional_records(root)
            functions, dependencies = sheet_formula_stats(records)
            all_functions.update(functions)
            all_dependencies[ws.title] = dependencies
            formula_records = [item for item in records if "formula_type" in item]
            literal_records = [item for item in records if "formula_type" not in item and item.get("value", "") != ""]
            merged = [str(item) for item in ws.merged_cells.ranges]
            hidden_rows = [str(index) for index, dim in ws.row_dimensions.items() if dim.hidden]
            hidden_cols = [key for key, dim in ws.column_dimensions.items() if dim.hidden]
            sheet_data.append({
                "ws": ws,
                "cached_ws": cached_wb[ws.title],
                "part": part,
                "records": records,
                "formulas": formula_records,
                "literals": literal_records,
                "validations": validations,
                "conditionals": conditionals,
                "merged": merged,
                "hidden_rows": hidden_rows,
                "hidden_cols": hidden_cols,
                "functions": functions,
                "dependencies": dependencies,
            })

        calc_pr = next(iter(workbook_root.xpath('//*[local-name()="calcPr"]')), None)
        defined = []
        for node in workbook_root.xpath('//*[local-name()="definedName"]'):
            defined.append({
                "name": node.get("name", ""),
                "localSheetId": node.get("localSheetId", ""),
                "hidden": node.get("hidden", ""),
                "value": text_of(node),
            })
        package_parts = zf.namelist()

    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    lines: list[str] = []
    add = lines.append
    add("# DND 5E 人物卡模板 XLSX 完整解析说明")
    add("")
    add("> 本文由 `tools/extract_xlsx_spec.py` 从 XLSX 原始 OOXML 与 OpenPyXL 对象模型联合生成。")
    add("> 公式以 `xl/worksheets/*.xml` 中的 `<f>` 为准；OpenPyXL 对本文件的扩展验证会发出兼容性警告，因此不能只依赖其高层读取结果。")
    add("")
    add("## 1. 文件指纹与工作簿概况")
    add("")
    add(f"- 源文件：`{source}`")
    add(f"- 文件大小：{source.stat().st_size:,} bytes")
    add(f"- SHA-256：`{digest}`")
    add(f"- 工作表数：{len(wb.worksheets)}")
    add(f"- OOXML 包部件数：{len(package_parts)}")
    add(f"- 命名范围数：{len(defined)}")
    add(f"- 计算模式：`{calc_pr.get('calcMode', 'auto') if calc_pr is not None else '未声明'}`")
    add(f"- 完整计算标记：`fullCalcOnLoad={calc_pr.get('fullCalcOnLoad', '') if calc_pr is not None else ''}`，`forceFullCalc={calc_pr.get('forceFullCalc', '') if calc_pr is not None else ''}`")
    add(f"- calcChain：{'存在' if 'xl/calcChain.xml' in package_parts else '不存在'}")
    add(f"- 外部链接：{len([x for x in package_parts if x.startswith('xl/externalLinks/')])}")
    add(f"- VBA：{'存在' if any('vbaProject' in x for x in package_parts) else '不存在'}")
    add(f"- 媒体文件：{', '.join(x for x in package_parts if x.startswith('xl/media/')) or '无'}")
    add("")
    add("| 工作表 | 状态 | 尺寸 | 原始单元格 | 字面值 | 单元格公式 | 合并区 | 数据验证 | 条件格式规则 | 隐藏行/列 |")
    add("|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|")
    for item in sheet_data:
        ws = item["ws"]
        add(f"| {esc(ws.title)} | {esc(ws.sheet_state)} | {ws.max_row}×{ws.max_column} | {len(item['records'])} | {len(item['literals'])} | {len(item['formulas'])} | {len(item['merged'])} | {len(item['validations'])} | {len(item['conditionals'])} | {len(item['hidden_rows'])}/{len(item['hidden_cols'])} |")

    add("")
    add("## 2. 关键结论")
    add("")
    total_formulas = sum(len(item["formulas"]) for item in sheet_data)
    add(f"1. 工作簿实际包含 **{total_formulas:,} 个单元格公式**。这些公式直接从 OOXML `<f>` 节点提取；不能以 OpenPyXL 高层 API 返回的公式数为准。")
    add("2. 模板同时使用普通数据验证与 Excel 扩展验证、共享公式、命名范围、隐藏辅助表和计算链。只复制可见四张人物卡页会破坏下拉与计算。")
    add("3. `图表`、`法术大全` 是规则/查找数据源，`BUY点计算器` 是购点辅助，`更新日志` 含版本说明；它们不是可随意删除的装饰页。")
    add("4. Python/OpenPyXL 只能保存公式文本，不能计算公式。写入后若需要新缓存值，必须由 Excel、LibreOffice 或其他公式引擎重算。")
    add("5. 本文件没有 VBA 和外部工作簿链接；公式依赖均位于本工作簿内部。")

    add("")
    add("## 3. 公式总体统计")
    add("")
    add("### 3.1 常用函数")
    add("")
    add("| 函数 | 出现次数 |")
    add("|---|---:|")
    for name, count in all_functions.most_common():
        add(f"| `{esc(name)}` | {count} |")
    add("")
    add("### 3.2 跨表依赖")
    add("")
    add("| 来源表 | 目标表 | 引用次数 |")
    add("|---|---|---:|")
    for source_name, deps in all_dependencies.items():
        for target, count in deps.most_common():
            add(f"| {esc(source_name)} | {esc(target)} | {count} |")

    add("")
    add("## 4. 工作表逐表规格")
    for item in sheet_data:
        ws = item["ws"]
        add("")
        add(f"### 4.{wb.sheetnames.index(ws.title) + 1} `{ws.title}`")
        add("")
        add(f"- OOXML 部件：`{item['part']}`")
        add(f"- 状态：`{ws.sheet_state}`；有效尺寸：`{ws.calculate_dimension()}`")
        add(f"- 冻结窗格：`{ws.freeze_panes or '无'}`；自动筛选：`{ws.auto_filter.ref or '无'}`")
        add(f"- 打印区域：`{ws.print_area or '无'}`；打印标题：`{ws.print_title_rows or ''} {ws.print_title_cols or ''}`")
        add(f"- 页面方向：`{ws.page_setup.orientation or '未设置'}`；纸张：`{ws.page_setup.paperSize or '未设置'}`；缩放：`{ws.page_setup.scale or '未设置'}`")
        add(f"- 隐藏行：{', '.join(item['hidden_rows']) or '无'}")
        add(f"- 隐藏列：{', '.join(item['hidden_cols']) or '无'}")
        add(f"- 公式：{len(item['formulas'])}；字面值：{len(item['literals'])}；合并区：{len(item['merged'])}；验证：{len(item['validations'])}")
        if item["dependencies"]:
            add("- 公式依赖：" + "；".join(f"`{name}` × {count}" for name, count in item["dependencies"].most_common()))
        if item["functions"]:
            add("- 主要函数：" + "；".join(f"`{name}` × {count}" for name, count in item["functions"].most_common(15)))

        add("")
        add("#### 全部公式")
        add("")
        if not item["formulas"]:
            add("无单元格公式。")
        else:
            add("| 单元格 | 类型/共享范围 | 公式 | 缓存值 | 样式 ID |")
            add("|---|---|---|---|---:|")
            for record in item["formulas"]:
                formula_meta = record.get("formula_type", "normal")
                if record.get("formula_ref"):
                    formula_meta += f" ref={record['formula_ref']}"
                add(f"| `{record['coord']}` | {esc(formula_meta)} | `{esc('=' + record.get('formula', ''))}` | {esc(record.get('cached', ''))} | {record['style']} |")

        add("")
        add("#### 全部字面单元格")
        add("")
        if not item["literals"]:
            add("无字面值。")
        else:
            add("| 单元格 | 类型 | 内容 | 样式 ID |")
            add("|---|---|---|---:|")
            for record in item["literals"]:
                add(f"| `{record['coord']}` | `{record['type']}` | {esc(record.get('value', ''))} | {record['style']} |")

        add("")
        add("#### 数据验证")
        add("")
        if not item["validations"]:
            add("无。")
        else:
            add("| 作用区域 | 类型 | 运算符 | 允许空白 | 公式/来源 | 提示与错误 |")
            add("|---|---|---|---|---|---|")
            for record in item["validations"]:
                messages = "；".join(x for x in [record["promptTitle"], record["prompt"], record["errorTitle"], record["error"]] if x)
                add(f"| `{esc(record['sqref'])}` | `{esc(record['type'])}` | `{esc(record['operator'])}` | `{esc(record['allowBlank'])}` | `{esc(record['formulas'])}` | {esc(messages)} |")

        add("")
        add("#### 条件格式")
        add("")
        if not item["conditionals"]:
            add("无。")
        else:
            add("| 作用区域 | 类型 | 优先级 | 运算符 | dxfId | 公式 |")
            add("|---|---|---:|---|---:|---|")
            for record in item["conditionals"]:
                add(f"| `{esc(record['sqref'])}` | `{esc(record['type'])}` | {esc(record['priority'])} | `{esc(record['operator'])}` | {esc(record['dxfId'])} | `{esc(record['formulas'])}` |")

        add("")
        add("#### 合并区域")
        add("")
        add(", ".join(f"`{value}`" for value in item["merged"]) if item["merged"] else "无。")

    add("")
    add("## 5. 命名范围完整清单")
    add("")
    add("命名范围是本模板下拉列表、种族/职业/背景配置、装备、法术与特性查找的核心 API。重命名或删除会使 `INDIRECT`、验证列表和查找公式失效。")
    add("")
    add("| 名称 | 局部表 ID | 隐藏 | 定义 |")
    add("|---|---:|---|---|")
    for item in defined:
        add(f"| `{esc(item['name'])}` | {esc(item['localSheetId'])} | {esc(item['hidden'])} | `{esc(item['value'])}` |")

    add("")
    add("## 6. 样式与版式索引")
    add("")
    style_usage = Counter(record["style"] for item in sheet_data for record in item["records"])
    add("| 样式 ID | 使用单元格数 | numFmtId | fontId | fillId | borderId | alignmentId | protectionId |")
    add("|---:|---:|---:|---:|---:|---:|---:|---:|")
    for style_id, count in sorted(style_usage.items(), key=lambda pair: int(pair[0])):
        try:
            style = wb._cell_styles[int(style_id)]
            add(f"| {style_id} | {count} | {style.numFmtId} | {style.fontId} | {style.fillId} | {style.borderId} | {style.alignmentId} | {style.protectionId} |")
        except (IndexError, ValueError):
            add(f"| {style_id} | {count} | ? | ? | ? | ? | ? | ? |")

    add("")
    add("## 7. 批注、图片与对象")
    add("")
    comments = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append((ws.title, cell.coordinate, cell.comment.author, cell.comment.text))
    add(f"- 批注数：{len(comments)}")
    add(f"- 图片数：{sum(len(ws._images) for ws in wb.worksheets)}")
    add(f"- 图表对象数：{sum(len(ws._charts) for ws in wb.worksheets)}")
    add("")
    if comments:
        add("| 工作表 | 单元格 | 作者 | 内容 |")
        add("|---|---|---|---|")
        for sheet, coord, author, value in comments:
            add(f"| {esc(sheet)} | `{coord}` | {esc(author)} | {esc(value)} |")

    add("")
    add("## 8. 程序化读写注意事项")
    add("")
    add("1. **公式源与缓存分离**：`<f>` 是公式，`<v>` 是上次计算缓存。解析器必须保存两者，不能把缓存当公式。")
    add("2. **共享公式**：报告已按主公式锚点展开共享公式；写回时可写普通公式，不必复原共享压缩。")
    add("3. **合并单元格**：写入合并区任意坐标前，应解析到合并区左上角。项目中的 `set_sheet_value()` 已实现这一点。")
    add("4. **扩展数据验证**：OpenPyXL 会警告并可能移除不支持的扩展。若只更新值，优先在原文件上做最小修改，并对输出包的验证节点做回归比较。")
    add("5. **重算**：OpenPyXL 不执行 Excel 公式。写入角色数据后应设置完整重算标记，并在需要缓存值时交给 Excel/LibreOffice 重算。")
    add("6. **命名范围是契约**：下拉和 `INDIRECT` 大量依赖中文名称。业务 JSON 到 Excel 的转换应引用既有名称，不要自行改名。")
    add("7. **隐藏辅助表不可删除**：`图表`、`法术大全`、购点计算器共同构成规则数据层。")
    add("8. **完整性回归**：建议至少比较工作表集合、公式坐标集合、命名范围集合、数据验证集合、合并区集合与关键输入/输出单元格。")

    add("")
    add("## 9. 与当前项目转换器的关系")
    add("")
    add("当前项目的 `backend/app/tools/character_builder.py` 已实现部分字段的 JSON → XLSX 与 XLSX → raw JSON 映射，但模板本身远比当前映射丰富。")
    add("本报告的逐单元格、逐公式、逐验证和逐命名范围清单可作为扩展转换器时的权威坐标目录。若要求真正无损往返，必须逐步覆盖背包、装备、职业/种族特性、法术位、全部法术行、背景下拉和辅助表引用，而不能只比较姓名、职业、等级与六维。")

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps({
        "source": str(source),
        "output": str(output),
        "sheets": len(sheet_data),
        "formulas": total_formulas,
        "defined_names": len(defined),
        "validations": sum(len(item["validations"]) for item in sheet_data),
        "cells": sum(len(item["records"]) for item in sheet_data),
        "output_bytes": output.stat().st_size,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
